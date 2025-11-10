from flask import Blueprint, render_template, request, session, jsonify, Response
from datetime import datetime, timedelta
from modules.auth.routes import login_required
from extensions import supabase_admin
import csv
import io
import json
import re
import os
import zipfile
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

# Blueprint acessível por todas as roles
export_relatorios_bp = Blueprint(
    'export_relatorios',
    __name__,
    url_prefix='/export_relatorios',
    template_folder='templates',
    static_folder='static',
    static_url_path='/export_relatorios/static'
)

@export_relatorios_bp.route('/')
@login_required
def index():
    """Tela inicial de exportação de relatórios (processos > 6 meses).
    Primeira etapa: apenas layout base e endpoint para futura busca.
    """
    user = session.get('user', {})
    user_companies = user.get('user_companies') or []
    user_role = user.get('role', '')
    
    # Para clientes, mostrar os CNPJs que eles têm acesso
    security_info = {
        'role': user_role,
        'companies': user_companies,
        'has_companies': len(user_companies) > 0
    }
    
    return render_template('export_relatorios/export_relatorios.html', 
                         user=user, 
                         security_info=security_info)

@export_relatorios_bp.route('/api/processos_antigos')
@login_required
def api_processos_antigos():
    """Endpoint inicial para futura extração de processos antigos (> 6 meses).
    Nesta primeira etapa retornamos apenas um payload mock para validar a página.
    """
    try:
        cutoff_date = (datetime.now() - timedelta(days=180)).strftime('%Y-%m-%d')
        return jsonify({
            'success': True,
            'message': 'Endpoint inicial pronto. Implementação de consulta será adicionada.',
            'cutoff_date': cutoff_date,
            'data': []
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ---------------------- NOVA IMPLEMENTAÇÃO DE BUSCA E EXPORTAÇÃO ----------------------

TABLE_COLUMNS = [
    'ref_unique','ref_importador','cnpj_importador','importador','modal','container','data_embarque',
    'data_chegada','transit_time_real','pais_procedencia','urf_despacho','exportador_fornecedor',
    'numero_di','data_registro','canal','data_desembaraco','mercadoria','data_abertura',
    'status_sistema','status_timeline','url_bandeira','despesas_processo','produtos_processo',
    'data_desova','limite_primeiro_periodo','limite_segundo_periodo','dias_extras_armazenagem',
    'valor_despesas_extras','documentos'
]

DATE_FIELDS = [c for c in TABLE_COLUMNS if c.startswith('data_')]


def serialize_cell_value(value):
    """Padroniza valores para exportação textual."""
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return value

def parse_br_date(date_str):
    """Converte data em formato DD/MM/YYYY para datetime; retorna None se inválida."""
    if not date_str or not isinstance(date_str, str):
        return None
    date_str = date_str.strip()
    # Aceitar já formato ISO
    try:
        if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            return datetime.strptime(date_str, '%Y-%m-%d')
    except Exception:
        pass
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except Exception:
        return None

def build_base_query(user):
    """Constrói query base com segurança obrigatória por CNPJ do usuário."""
    # Utiliza view consolidada com colunas atualizadas de status
    q = supabase_admin.table('vw_importacoes_geral_export').select('*')
    
    # SEGURANÇA OBRIGATÓRIA: Sempre filtrar por CNPJs do usuário
    user_role = user.get('role', '')
    user_companies = user.get('user_companies') or []
    perfil_principal = user.get('perfil_principal', '')
    is_admin_operacao = perfil_principal == 'admin_operacao'
    
    # Para clientes, SEMPRE restringir aos CNPJs associados
    if user_role == 'cliente_unique':
        if user_companies:
            print(f"[EXPORT_REL][SECURITY] Cliente {user.get('id')} restrito aos CNPJs: {user_companies}")
            q = q.in_('cnpj_importador', user_companies)
        else:
            # Se não tem empresas associadas, retorna query vazia por segurança
            print(f"[EXPORT_REL][SECURITY] Cliente {user.get('id')} sem CNPJs - bloqueando acesso")
            q = q.limit(0)
    
    # Para usuários internos com perfil admin_operacao, permitir ver todos os dados
    elif user_role == 'interno_unique' and is_admin_operacao:
        print(f"[EXPORT_REL][SECURITY] Usuário admin_operacao {user.get('id')} - acesso completo a todos os dados")
    
    # Para usuários internos normais, aplicar filtro se tiverem CNPJs específicos
    elif user_role == 'interno_unique' and user_companies:
        print(f"[EXPORT_REL][SECURITY] Usuário interno {user.get('id')} filtrado por CNPJs: {user_companies}")
        q = q.in_('cnpj_importador', user_companies)
    
    # Admins podem acessar todos os dados (sem filtro adicional)
    elif user_role == 'admin':
        print(f"[EXPORT_REL][SECURITY] Admin {user.get('id')} - acesso completo")
    
    else:
        # Usuários sem role definida ou roles desconhecidas = acesso negado
        print(f"[EXPORT_REL][SECURITY] Usuário {user.get('id')} com role '{user_role}' - acesso negado")
        q = q.limit(0)
    
    return q

def apply_query_filters(q, filters: dict, user: dict):
    """Aplica filtros adicionais, mas mantém a segurança de CNPJs do usuário."""
    import re
    user_role = user.get('role', '')
    user_companies = user.get('user_companies') or []
    
    # Campos que suportam busca múltipla (separados por vírgula)
    multi_search_fields = ['ref_importador', 'ref_unique', 'numero_di', 'container']
    
    # Campos de data range (não processar no server-side, serão filtrados em Python)
    date_range_fields = [
        'data_embarque_start', 'data_embarque_end',
        'data_chegada_start', 'data_chegada_end', 
        'data_registro_start', 'data_registro_end',
        'data_desembaraco_start', 'data_desembaraco_end'
    ]
    
    # Campos que faremos server-side (igualdade ou like simples)
    for col, val in filters.items():
        if col in ['date_start','date_end','older_than_days','page','page_size','export'] + date_range_fields:
            continue
        if val is None or val == '':
            continue
            
        # SEGURANÇA CRÍTICA: Ignorar filtro de cnpj_importador do usuário
        # A segurança já foi aplicada em build_base_query()
        if col == 'cnpj_importador':
            print(f"[EXPORT_REL][SECURITY] Ignorando filtro cnpj_importador do usuário {user.get('id')} - segurança já aplicada")
            continue
        
        # Tratamento para campos de busca múltipla
        if col in multi_search_fields and isinstance(val, str):
            # Separar por vírgula e limpar espaços
            values = [v.strip() for v in val.split(',') if v.strip()]
            if len(values) > 1:
                # Múltiplos valores: usar filtro OR com in_()
                print(f"[EXPORT_REL][MULTI_SEARCH] Campo {col} com {len(values)} valores: {values}")
                q = q.in_(col, values)
                continue
            elif len(values) == 1:
                # Um único valor após limpeza
                val = values[0]
            
        # Tratamento numérico
        if col in ['transit_time_real']:
            try:
                float(val)  # valida
                q = q.eq(col, val)
            except ValueError:
                pass
        else:
            # Usar ilike para texto quando cabível (evitar wildcard se usuário não fornecer)
            if isinstance(val, str) and ('%' in val or len(val) > 3):
                q = q.ilike(col, val if '%' in val else f"%{val}%")
            else:
                q = q.eq(col, val)
    return q

def validate_user_data_access(rows, user):
    """Validação adicional: verifica se todos os registros pertencem aos CNPJs do usuário.
    Esta é uma camada extra de segurança contra bypass de API.
    """
    user_role = user.get('role', '')
    user_companies = user.get('user_companies') or []
    perfil_principal = user.get('perfil_principal', '')
    is_admin_operacao = perfil_principal == 'admin_operacao'
    
    # Para clientes, verificar se TODOS os registros pertencem aos CNPJs permitidos
    if user_role == 'cliente_unique':
        if not user_companies:
            print(f"[EXPORT_REL][SECURITY_VALIDATION] Cliente {user.get('id')} sem CNPJs - removendo todos os registros")
            return []
        
        valid_rows = []
        blocked_count = 0
        
        for row in rows:
            row_cnpj = row.get('cnpj_importador', '')
            if row_cnpj in user_companies:
                valid_rows.append(row)
            else:
                blocked_count += 1
                print(f"[EXPORT_REL][SECURITY_VALIDATION] Bloqueado acesso ao CNPJ {row_cnpj} para cliente {user.get('id')}")
        
        if blocked_count > 0:
            print(f"[EXPORT_REL][SECURITY_VALIDATION] Total de {blocked_count} registros bloqueados por segurança")
        
        return valid_rows
    
    # Para usuários internos com perfil admin_operacao, permitir ver todos os dados
    elif user_role == 'interno_unique' and is_admin_operacao:
        print(f"[EXPORT_REL][SECURITY_VALIDATION] Usuário admin_operacao {user.get('id')} - acesso a todos os registros")
        return rows
    
    # Para usuários internos com CNPJs específicos, aplicar mesma validação
    elif user_role == 'interno_unique' and user_companies:
        valid_rows = []
        for row in rows:
            row_cnpj = row.get('cnpj_importador', '')
            if row_cnpj in user_companies:
                valid_rows.append(row)
        return valid_rows
    
    # Admins podem ver todos os dados
    elif user_role == 'admin':
        return rows
    
    # Outras roles = sem acesso
    else:
        print(f"[EXPORT_REL][SECURITY_VALIDATION] Role '{user_role}' não autorizada - removendo todos os registros")
        return []

def post_fetch_filter(rows, filters):
    """Filtragem em Python para datas (formato texto). Suporta múltiplos intervalos de data."""
    date_start = filters.get('date_start')
    date_end = filters.get('date_end')
    
    # Filtros de data range específicos
    data_embarque_start = filters.get('data_embarque_start')
    data_embarque_end = filters.get('data_embarque_end')
    data_chegada_start = filters.get('data_chegada_start')
    data_chegada_end = filters.get('data_chegada_end')
    data_registro_start = filters.get('data_registro_start')
    data_registro_end = filters.get('data_registro_end')
    data_desembaraco_start = filters.get('data_desembaraco_start')
    data_desembaraco_end = filters.get('data_desembaraco_end')
    
    # Identificadores que devem permitir ignorar o recorte padrão de 6 meses quando usados
    identifier_filters = ['ref_unique','numero_di','ref_importador','container']
    has_identifier = any(filters.get(f) for f in identifier_filters)

    # Parse das datas principais
    dt_start = parse_br_date(date_start) if date_start else None
    dt_end = parse_br_date(date_end) if date_end else None
    
    # Parse dos filtros de data range
    dt_embarque_start = parse_br_date(data_embarque_start) if data_embarque_start else None
    dt_embarque_end = parse_br_date(data_embarque_end) if data_embarque_end else None
    dt_chegada_start = parse_br_date(data_chegada_start) if data_chegada_start else None
    dt_chegada_end = parse_br_date(data_chegada_end) if data_chegada_end else None
    dt_registro_start = parse_br_date(data_registro_start) if data_registro_start else None
    dt_registro_end = parse_br_date(data_registro_end) if data_registro_end else None
    dt_desembaraco_start = parse_br_date(data_desembaraco_start) if data_desembaraco_start else None
    dt_desembaraco_end = parse_br_date(data_desembaraco_end) if data_desembaraco_end else None
    
    cutoff = None  # Sem cutoff automático agora

    def row_pass(r):
        # Filtro principal (data_abertura/data_registro)
        raw = r.get('data_abertura') or r.get('data_registro')
        dt = parse_br_date(raw)
        if cutoff and dt:
            if dt > cutoff:
                return False
        if dt_start and dt and dt < dt_start:
            return False
        if dt_end and dt and dt > dt_end:
            return False
        
        # Filtro de data embarque
        if dt_embarque_start or dt_embarque_end:
            dt_embarque = parse_br_date(r.get('data_embarque'))
            if dt_embarque:
                if dt_embarque_start and dt_embarque < dt_embarque_start:
                    return False
                if dt_embarque_end and dt_embarque > dt_embarque_end:
                    return False
        
        # Filtro de data chegada
        if dt_chegada_start or dt_chegada_end:
            dt_chegada = parse_br_date(r.get('data_chegada'))
            if dt_chegada:
                if dt_chegada_start and dt_chegada < dt_chegada_start:
                    return False
                if dt_chegada_end and dt_chegada > dt_chegada_end:
                    return False
        
        # Filtro de data registro
        if dt_registro_start or dt_registro_end:
            dt_reg = parse_br_date(r.get('data_registro'))
            if dt_reg:
                if dt_registro_start and dt_reg < dt_registro_start:
                    return False
                if dt_registro_end and dt_reg > dt_registro_end:
                    return False
        
        # Filtro de data desembaraço
        if dt_desembaraco_start or dt_desembaraco_end:
            dt_desemb = parse_br_date(r.get('data_desembaraco'))
            if dt_desemb:
                if dt_desembaraco_start and dt_desemb < dt_desembaraco_start:
                    return False
                if dt_desembaraco_end and dt_desemb > dt_desembaraco_end:
                    return False
        
        return True

    if dt_start or dt_end or dt_embarque_start or dt_embarque_end or dt_chegada_start or dt_chegada_end or dt_registro_start or dt_registro_end or dt_desembaraco_start or dt_desembaraco_end:
        before = len(rows)
        rows = [r for r in rows if row_pass(r)]
        after = len(rows)
        print(f"[EXPORT_REL][FILTRO_DATAS] has_identifier={has_identifier} dt_start={dt_start} dt_end={dt_end} antes={before} depois={after}")
    return rows

def paginate(rows, page, page_size):
    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size
    return rows[start:end], total

def get_documentos_by_ref_unique(ref_unique_list):
    """
    Busca documentos ativos para uma lista de ref_unique.
    Retorna dict: {ref_unique: [documentos]}
    """
    if not ref_unique_list:
        return {}
    
    try:
        # Buscar documentos ativos para os processos (incluindo descrição)
        response = supabase_admin.table('documentos_processos')\
            .select('ref_unique, nome_exibicao, storage_path, extensao, tamanho_bytes, data_upload, descricao')\
            .in_('ref_unique', ref_unique_list)\
            .eq('ativo', True)\
            .order('data_upload', desc=True)\
            .execute()
        
        documentos = response.data or []
        
        # Agrupar por ref_unique
        docs_by_ref = {}
        for doc in documentos:
            ref = doc.get('ref_unique')
            if ref not in docs_by_ref:
                docs_by_ref[ref] = []
            
            # Gerar URL pública do documento
            storage_path = doc.get('storage_path')
            doc_url = None
            if storage_path:
                try:
                    # Gerar URL assinada com 1 hora de validade
                    url_response = supabase_admin.storage.from_('processos-documentos').create_signed_url(storage_path, 3600)
                    doc_url = url_response.get('signedURL')
                except Exception as url_error:
                    print(f"[EXPORT_REL][DOCS] Erro ao gerar URL para {storage_path}: {url_error}")
            
            docs_by_ref[ref].append({
                'nome': doc.get('nome_exibicao'),
                'extensao': doc.get('extensao'),
                'tamanho': doc.get('tamanho_bytes'),
                'data_upload': doc.get('data_upload'),
                'descricao': doc.get('descricao'),  # NOVO: incluir descrição
                'url': doc_url
            })
        
        print(f"[EXPORT_REL][DOCS] Encontrados documentos para {len(docs_by_ref)} processos")
        return docs_by_ref
        
    except Exception as e:
        print(f"[EXPORT_REL][DOCS][ERRO] {e}")
        import traceback
        traceback.print_exc()
        return {}

def extract_filters(req_json):
    filters = {}
    for col in TABLE_COLUMNS:
        if col in req_json:
            filters[col] = req_json.get(col)
    # Extras
    for extra in ['date_start','date_end','older_than_days','page','page_size','export']:
        if extra in req_json:
            filters[extra] = req_json.get(extra)
    return filters

@export_relatorios_bp.route('/api/search_processos', methods=['POST'])
def search_processos():
    """
    Executa busca com filtros e retorna JSON paginado.
    A coluna 'documentos' é buscada sob demanda quando necessária.
    """
    # Verificar bypass key ou sessão
    api_bypass_key = os.getenv('API_BYPASS_KEY')
    request_api_key = request.headers.get('X-API-Key')
    
    if not (api_bypass_key and request_api_key == api_bypass_key):
        # Validar sessão normal
        if 'user' not in session:
            return jsonify({'success': False, 'error': 'Não autenticado'}), 401
    
    started_at = datetime.now()
    user = session.get('user', {}) if 'user' in session else {'role': 'admin', 'user_companies': []}
    payload = request.get_json(silent=True) or {}
    filters = extract_filters(payload)
    page = int(filters.get('page') or 1)
    page_size = min(int(filters.get('page_size') or 500), 5000)
    print(f"[EXPORT_REL] Busca iniciada user={user.get('id')} role={user.get('role')} page={page} page_size={page_size}")
    try:
        q = build_base_query(user)
        q = apply_query_filters(q, filters, user)
        # Limite temporário grande para pós-filtragem; evitar extrair tudo indiscriminadamente
        raw = q.limit(20000).execute()
        rows = raw.data or []
        print(f"[EXPORT_REL] Registros retornados antes pós-filtro: {len(rows)}")
        
        # VALIDAÇÃO DE SEGURANÇA: Verificar se todos os registros pertencem ao usuário
        rows = validate_user_data_access(rows, user)
        print(f"[EXPORT_REL] Após validação de segurança: {len(rows)}")
        
        rows = post_fetch_filter(rows, filters)
        print(f"[EXPORT_REL] Após pós-filtro: {len(rows)}")
        
        # OTIMIZAÇÃO: Buscar documentos APENAS para a página atual
        page_rows, total = paginate(rows, page, page_size)
        
        if page_rows:
            ref_unique_list = [r.get('ref_unique') for r in page_rows if r.get('ref_unique')]
            print(f"[EXPORT_REL] Buscando documentos para {len(ref_unique_list)} processos na página")
            documentos_map = get_documentos_by_ref_unique(ref_unique_list)
            
            # Adicionar documentos aos registros
            for row in page_rows:
                ref = row.get('ref_unique')
                row['documentos'] = documentos_map.get(ref, [])
        
        duration = (datetime.now() - started_at).total_seconds()
        return jsonify({
            'success': True,
            'duration': duration,
            'page': page,
            'page_size': page_size,
            'returned': len(page_rows),
            'total_count': total,
            'columns': TABLE_COLUMNS,
            'rows': page_rows
        })
    except Exception as e:
        duration = (datetime.now() - started_at).total_seconds()
        print(f"[EXPORT_REL][ERRO] Erro após {duration:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_relatorios_bp.route('/api/export_csv', methods=['POST'])
@login_required
def export_csv():
    """
    Exporta CSV com os filtros informados (limite de segurança).
    Otimizado para lidar com grandes volumes (20k+) de registros.
    """
    started_at = datetime.now()
    user = session.get('user', {})
    payload = request.get_json(silent=True) or {}
    filters = extract_filters(payload)
    max_rows = 100000  # Aumentado de 50k para 100k
    
    print(f"[EXPORT_REL] Export CSV iniciado user={user.get('id')} filtros={filters}")
    try:
        q = build_base_query(user)
        q = apply_query_filters(q, filters, user)
        
        # OTIMIZAÇÃO: Buscar sem a coluna 'documentos' para ganhar performance
        # A busca de documentos é cara e não será incluída na exportação
        print(f"[EXPORT_REL] Buscando até {max_rows} registros (documentos excluídos para performance)")
        raw = q.limit(max_rows + 1).execute()
        rows = raw.data or []
        
        # VALIDAÇÃO DE SEGURANÇA: Verificar se todos os registros pertencem ao usuário
        rows = validate_user_data_access(rows, user)
        print(f"[EXPORT_REL] Após validação de segurança: {len(rows)}")
        
        rows = post_fetch_filter(rows, filters)
        
        # Validar limite e avisar se foi truncado
        if len(rows) > max_rows:
            print(f"[EXPORT_REL] AVISO: Resultado truncado de {len(rows)} para {max_rows}")
            rows = rows[:max_rows]
        
        # Colunas a usar (documentos é buscado sob demanda na página)
        columns_to_export = [c for c in TABLE_COLUMNS if c != 'documentos']
        
        # Gerar CSV em stream
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Usar colunas sem documentos
        writer.writerow(columns_to_export)
        
        for r in rows:
            writer.writerow([serialize_cell_value(r.get(col)) for col in columns_to_export])
        
        csv_data = output.getvalue()
        output.close()
        
        duration = (datetime.now() - started_at).total_seconds()
        filename = f"export_processos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        headers = {
            'Content-Disposition': f'attachment; filename={filename}',
            'X-Export-Duration': f"{duration:.2f}s",
            'X-Export-Rows': str(len(rows))
        }
        print(f"[EXPORT_REL] CSV gerado com {len(rows)} registros em {duration:.2f}s para usuário {user.get('id')}")
        return Response(csv_data, mimetype='text/csv; charset=utf-8', headers=headers)
    except Exception as e:
        duration = (datetime.now() - started_at).total_seconds()
        print(f"[EXPORT_REL][ERRO_EXPORT_CSV] Erro após {duration:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e), 'duration': f"{duration:.2f}s"}), 500

@export_relatorios_bp.route('/api/export_excel', methods=['POST'])
@login_required
def export_excel():
    """
    Exporta Excel (XLSX) com os filtros informados (limite de segurança).
    Otimizado para lidar com grandes volumes (20k+) de registros.
    """
    started_at = datetime.now()
    user = session.get('user', {})
    payload = request.get_json(silent=True) or {}
    filters = extract_filters(payload)
    max_rows = 100000  # Aumentado de 50k para 100k
    
    print(f"[EXPORT_REL] Export Excel iniciado user={user.get('id')} filtros={filters}")
    try:
        q = build_base_query(user)
        q = apply_query_filters(q, filters, user)
        
        # OTIMIZAÇÃO: Buscar sem a coluna 'documentos' para ganhar performance
        # A busca de documentos é cara e não será incluída na exportação
        print(f"[EXPORT_REL] Buscando até {max_rows} registros (documentos excluídos para performance)")
        raw = q.limit(max_rows + 1).execute()
        rows = raw.data or []
        
        # VALIDAÇÃO DE SEGURANÇA: Verificar se todos os registros pertencem ao usuário
        rows = validate_user_data_access(rows, user)
        print(f"[EXPORT_REL] Após validação de segurança: {len(rows)}")
        
        rows = post_fetch_filter(rows, filters)
        
        # Validar limite e avisar se foi truncado
        if len(rows) > max_rows:
            print(f"[EXPORT_REL] AVISO: Resultado truncado de {len(rows)} para {max_rows}")
            rows = rows[:max_rows]
        
        # Colunas a usar (documentos é buscado sob demanda na página)
        columns_to_export = [c for c in TABLE_COLUMNS if c != 'documentos']
        
        print(f"[EXPORT_REL] Gerando Excel com {len(rows)} registros")
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Processos"
        
        # Estilização do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escrever cabeçalho
        for col_idx, col_name in enumerate(columns_to_export, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escrever dados em batch para otimizar
        print(f"[EXPORT_REL] Escrevendo dados no Excel...")
        for row_idx, row_data in enumerate(rows, 2):
            if row_idx % 5000 == 0:
                print(f"[EXPORT_REL] Progresso: {row_idx}/{len(rows)} registros")
            
            for col_idx, col_name in enumerate(columns_to_export, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = serialize_cell_value(row_data.get(col_name))
        
        # Ajustar largura das colunas
        for col_idx in range(1, len(columns_to_export) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Salvar em BytesIO
        print(f"[EXPORT_REL] Salvando arquivo Excel...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        duration = (datetime.now() - started_at).total_seconds()
        filename = f"export_processos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename={filename}',
            'X-Export-Duration': f"{duration:.2f}s",
            'X-Export-Rows': str(len(rows))
        }
        print(f"[EXPORT_REL] Excel gerado com {len(rows)} registros em {duration:.2f}s para usuário {user.get('id')}")
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    except Exception as e:
        duration = (datetime.now() - started_at).total_seconds()
        print(f"[EXPORT_REL][ERRO_EXPORT_EXCEL] Erro após {duration:.2f}s: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e), 'duration': f"{duration:.2f}s"}), 500

@export_relatorios_bp.route('/api/filter_options', methods=['GET'])
def get_filter_options():
    """
    Retorna opções disponíveis para campos categóricos (dropdowns/checkboxes).
    Usa cache e limita quantidade para performance.
    """
    # Verificar bypass key ou sessão
    api_bypass_key = os.getenv('API_BYPASS_KEY')
    request_api_key = request.headers.get('X-API-Key')
    
    if not (api_bypass_key and request_api_key == api_bypass_key):
        # Validar sessão normal
        if 'user' not in session:
            return jsonify({'success': False, 'error': 'Não autenticado'}), 401
    
    try:
        user = session.get('user', {}) if 'user' in session else {'role': 'admin', 'user_companies': []}
        
        # Campos categóricos que vamos buscar opções
        categorical_fields = {
            'modal': {'limit': 10},
            'canal': {'limit': 10},
            'status_sistema': {'limit': 30},
            'status_timeline': {'limit': 30},
            'pais_procedencia': {'limit': 50},
            'urf_despacho': {'limit': 50},
            'mercadoria': {'limit': 100}
        }
        
        result = {}
        
        for field, config in categorical_fields.items():
            # Buscar todos os valores e filtrar em Python (mais simples para valores distintos)
            q = build_base_query(user)
            
            # Limitar a busca inicial
            response = q.limit(5000).execute()
            
            # Extrair valores únicos
            values = []
            seen = set()
            for row in response.data or []:
                val = row.get(field)
                if val and val != 'null' and val not in seen:
                    values.append(val)
                    seen.add(val)
                    if len(values) >= config['limit']:
                        break
            
            # Ordenar os valores
            values.sort()
            
            result[field] = {
                'values': values[:config['limit']],
                'count': len(values),
                'limited': len(values) >= config['limit']
            }
        
        print(f"[EXPORT_REL][FILTER_OPTIONS] Retornando opções para {len(result)} campos")
        
        return jsonify(result)
        
    except Exception as e:
        print(f"[EXPORT_REL][ERRO_FILTER_OPTIONS] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@export_relatorios_bp.route('/api/download_all_docs/<path:ref_unique>', methods=['GET'])
def download_all_docs(ref_unique):
    """
    Baixa todos os documentos de um processo como arquivo ZIP.
    Usa <path:ref_unique> para aceitar ref_unique com barra (ex: UN25/6564)
    """
    # Verificar bypass key ou sessão
    api_bypass_key = os.getenv('API_BYPASS_KEY')
    request_api_key = request.headers.get('X-API-Key')
    
    if not (api_bypass_key and request_api_key == api_bypass_key):
        # Validar sessão normal
        if 'user' not in session:
            return jsonify({'success': False, 'error': 'Não autenticado'}), 401
    
    try:
        print(f"[EXPORT_REL][ZIP] Iniciando download de documentos para {ref_unique}")
        
        # Buscar documentos do processo (incluindo descrição)
        response = supabase_admin.table('documentos_processos')\
            .select('nome_exibicao, storage_path, extensao, descricao')\
            .eq('ref_unique', ref_unique)\
            .eq('ativo', True)\
            .execute()
        
        documentos = response.data or []
        
        if not documentos:
            return jsonify({'success': False, 'error': 'Nenhum documento encontrado'}), 404
        
        print(f"[EXPORT_REL][ZIP] Encontrados {len(documentos)} documentos")
        
        # Criar ZIP em memória
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for doc in documentos:
                storage_path = doc.get('storage_path')
                nome_arquivo = doc.get('nome_exibicao')
                
                try:
                    # Gerar URL assinada
                    url_response = supabase_admin.storage.from_('processos-documentos').create_signed_url(storage_path, 300)
                    doc_url = url_response.get('signedURL')
                    
                    if not doc_url:
                        print(f"[EXPORT_REL][ZIP] Erro ao gerar URL para {storage_path}")
                        continue
                    
                    # Baixar arquivo do Supabase
                    doc_response = requests.get(doc_url, timeout=30)
                    
                    if doc_response.status_code == 200:
                        # Adicionar ao ZIP
                        zip_file.writestr(nome_arquivo, doc_response.content)
                        print(f"[EXPORT_REL][ZIP] Adicionado: {nome_arquivo}")
                    else:
                        print(f"[EXPORT_REL][ZIP] Erro ao baixar {nome_arquivo}: Status {doc_response.status_code}")
                        
                except Exception as doc_error:
                    print(f"[EXPORT_REL][ZIP] Erro ao processar {nome_arquivo}: {doc_error}")
                    continue
        
        # Preparar resposta
        zip_buffer.seek(0)
        zip_filename = f"{ref_unique.replace('/', '-')}_documentos.zip"
        
        print(f"[EXPORT_REL][ZIP] ZIP criado com sucesso: {zip_filename}")
        
        return Response(
            zip_buffer.getvalue(),
            mimetype='application/zip',
            headers={
                'Content-Disposition': f'attachment; filename={zip_filename}',
                'Content-Type': 'application/zip'
            }
        )
        
    except Exception as e:
        print(f"[EXPORT_REL][ZIP][ERRO] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
